from openpyxl import load_workbook
import pandas as pd
from sklearn.impute import KNNImputer

RED = 'FFFFCCCC'
PURPLE = 'FF6A5ACD'
YELLOW = 'FFFFFF33'
GRAY = 'FF808080'

NUM_SAMPLES = 323



def read_data(file_path):
    wb = load_workbook(file_path)
    ws = wb['NMR']
    
    data = pd.DataFrame(ws.values)
    data = data.iloc[:NUM_SAMPLES + 1, :]
    data.columns = data.iloc[1]
    data = data.drop([0, 1])
    data = data.reset_index(drop=True)

    # Use openpyxl to get the color of each cell
    # Colors same shape as data
    colors = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        row_colors = []
        for cell in row:
            row_colors.append(cell.fill.start_color.index)
        colors.append(row_colors)
    colors.pop(0)
    colors = pd.DataFrame(colors)
    colors.columns = data.columns
    colors = colors.reset_index(drop=True)

    return data, colors

def drop_purple(data, colors):
    # Drop columns from the data if all of the following conditions are met:
    # 1. More than 50% of the cells are purple
    # 2. Within each group, more than 50% of the cells are purple
    # 3. The proportion of purple cells in each group are within 10% of each other

    # Returns: list of column names to keep
    PD_indices = data[data['ENROLLMENT_CATEGORY'] == 'PD'].index
    NON_PD_indices = data[data['ENROLLMENT_CATEGORY'] == 'No PD'].index

    # Iterate through each column
    drop_cols = []

    for i in range(colors.shape[1] ):
        col = colors.iloc[:, i]

        PD_purple = col[PD_indices] == PURPLE
        NON_PD_purple = col[NON_PD_indices] == PURPLE

        # Check if the column should be dropped
        if (PD_purple.mean() > 0.5) and (NON_PD_purple.mean() > 0.5):
            if (abs(PD_purple.mean() - NON_PD_purple.mean()) < 0.1):
                drop_cols.append(i)

    data[colors == PURPLE] = None

    return drop_cols


def replace_gray(data, colors):
    # Replace gray cells with the maximum value in the column

    for i in range(colors.shape[1]):
        col = colors.iloc[:, i]
        gray_indices = col[col == GRAY].index
        non_gray_yellow_purple = data.iloc[:, i][col != GRAY][col != YELLOW][col != PURPLE]
        data.iloc[gray_indices, i] = non_gray_yellow_purple.max()

    return data

def replace_yellow(data, colors):
    # Replace yellow cells with the median value among non-yellow and purple cells in the column

    for i in range(colors.shape[1]):
        col = colors.iloc[:, i]
        yellow_indices = col[col == YELLOW].index
        non_yellow_purple = data.iloc[:, i][col != YELLOW][col != PURPLE]
        data.iloc[yellow_indices, i] = non_yellow_purple.median()

    return data


def impute_purple(data, colors):
    # Impute purple cells using KNN imputation

    imputer = KNNImputer(n_neighbors=3)

    # Change all purple cells to NaN
    # data[colors == PURPLE] = None

    # Impute the data
    data = imputer.fit_transform(data)

    return data

def replace_NA(data):
    data = data.replace("NA", None) 
    data = data.replace("∞", None)   
    return data

if __name__ == '__main__':
    file_path = 'PD_Serum_Metabolomics_Final_NMR.xlsx'
    data, colors = read_data(file_path)

    drop_cols = drop_purple(data, colors)
    data = data.drop(data.columns[drop_cols], axis=1)
    colors = colors.drop(colors.columns[drop_cols], axis=1)
    labels = data['ENROLLMENT_CATEGORY']
    # Change labels so that they are 0, 1
    labels = labels.replace('PD', 1)
    labels = labels.replace('No PD', 0)

    # Additionally drop first five columns
    data = data.drop(data.columns[:5], axis=1)
    colors = colors.drop(colors.columns[:5], axis=1)

    data = replace_NA(data)
    data = replace_gray(data, colors)
    data = replace_yellow(data, colors)

    col_names = data.columns
    data = impute_purple(data, colors)
    

    df = pd.DataFrame(data, columns=col_names)
    df.insert(0, 'ENROLLMENT_CATEGORY', labels)
    df.to_csv('preprocessed_data_PD.csv', index=False)

    print('Preprocessing complete. Data saved to preprocessed_data_PD.csv')
